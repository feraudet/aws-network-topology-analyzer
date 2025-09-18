"""
Report Generator
Generates reports in multiple formats (JSON, CSV, Excel, HTML)
"""

import json
import logging
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional
from jinja2 import Template, Environment, FileSystemLoader
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from aws_network_discovery.config.settings import Config


logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generates comprehensive reports in multiple formats"""
    
    def __init__(self, config: Config):
        """
        Initialize report generator
        
        Args:
            config: Configuration object
        """
        self.config = config
        
    def generate_all_reports(self, analysis_results: Dict[str, Any], output_dir: str) -> None:
        """
        Generate all report formats
        
        Args:
            analysis_results: Analysis results from NetworkAnalyzer
            output_dir: Output directory for reports
        """
        logger.info(f"Generating reports in directory: {output_dir}")
        
        # Ensure output directory exists
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        try:
            # Generate JSON report
            if self.config.output.include_resource_inventory:
                logger.info("Generating JSON report...")
                self._generate_json_report(analysis_results, output_dir)
            
            # Generate CSV reports
            logger.info("Generating CSV reports...")
            self._generate_csv_reports(analysis_results, output_dir)
            
            # Generate Excel report
            logger.info("Generating Excel report...")
            self._generate_excel_report(analysis_results, output_dir)
            
            # Generate HTML report
            logger.info("Generating HTML report...")
            self._generate_html_report(analysis_results, output_dir)
            
            logger.info("All reports generated successfully")
            
        except Exception as e:
            logger.error(f"Failed to generate reports: {str(e)}")
            raise
    
    def _generate_json_report(self, analysis_results: Dict[str, Any], output_dir: str) -> None:
        """Generate comprehensive JSON report"""
        json_file = Path(output_dir) / "network_analysis_report.json"
        
        with open(json_file, 'w') as f:
            json.dump(
                analysis_results,
                f,
                indent=self.config.output.json_indent,
                default=str
            )
        
        logger.info(f"JSON report saved to {json_file}")
    
    def _generate_csv_reports(self, analysis_results: Dict[str, Any], output_dir: str) -> None:
        """Generate CSV reports for different data sections"""
        
        # Communication Paths CSV
        if analysis_results.get('communication_paths'):
            self._generate_communication_paths_csv(analysis_results['communication_paths'], output_dir)
        
        # Resource Inventory CSV
        if analysis_results.get('resource_inventory'):
            self._generate_resource_inventory_csv(analysis_results['resource_inventory'], output_dir)
        
        # Security Analysis CSV
        if analysis_results.get('security_analysis'):
            self._generate_security_analysis_csv(analysis_results['security_analysis'], output_dir)
    
    def _generate_communication_paths_csv(self, paths: List[Dict], output_dir: str) -> None:
        """Generate CSV for communication paths"""
        csv_file = Path(output_dir) / "communication_paths.csv"
        
        # Flatten path data for CSV
        flattened_paths = []
        for path in paths:
            flat_path = {
                'Source_Type': path['source']['type'],
                'Source_ID': path['source']['id'],
                'Source_Name': path['source']['name'],
                'Source_Region': path['source']['region'],
                'Source_Profile': path['source'].get('profile'),
                'Source_Account': path['source']['account_id'],
                'Source_Arn': path['source'].get('arn'),
                'Destination_Type': path['destination']['type'],
                'Destination_ID': path['destination']['id'],
                'Destination_Name': path['destination']['name'],
                'Destination_Region': path['destination']['region'],
                'Destination_Profile': path['destination'].get('profile'),
                'Destination_Account': path['destination']['account_id'],
                'Destination_Arn': path['destination'].get('arn'),
                'Protocol': path['protocol'],
                'Port_Range': path['port_range'],
                'Direction': path['direction'],
                'Is_Cross_Account': path['is_cross_account'],
                'Is_Cross_Region': path['is_cross_region'],
                'Is_Third_Party': path['is_third_party'],
                'Confidence_Score': path['confidence_score'],
                'Security_Groups_Valid': path['validation_results'].get('security_groups_valid', 'Unknown'),
                'NACLs_Valid': path['validation_results'].get('nacls_valid', 'Unknown'),
                'Route_Tables_Valid': path['validation_results'].get('route_tables_valid', 'Unknown'),
                'Firewall_Rules_Valid': path['validation_results'].get('firewall_rules_valid', 'Unknown'),
                'Path_Chain': ' -> '.join([step['description'] for step in path['path_chain']]),
            }
            flattened_paths.append(flat_path)
        
        # Create DataFrame and save to CSV
        df = pd.DataFrame(flattened_paths)
        df.to_csv(csv_file, index=False, sep=self.config.output.csv_delimiter)
        
        logger.info(f"Communication paths CSV saved to {csv_file}")
    
    def _generate_resource_inventory_csv(self, inventory: Dict[str, Any], output_dir: str) -> None:
        """Generate CSV for resource inventory"""
        
        # Compute resources
        if inventory.get('compute_resources'):
            csv_file = Path(output_dir) / "compute_resources.csv"
            df = pd.DataFrame(inventory['compute_resources'])
            # Ensure common columns
            for col in ['profile', 'account_id', 'arn']:
                if col not in df.columns:
                    df[col] = None
            # Add friendly kind for ELB
            if 'type' in df.columns:
                def _kind(row):
                    if row.get('type') == 'ELBv2LoadBalancer':
                        t = row.get('lb_type')
                        if t == 'application':
                            return 'ALB'
                        if t == 'network':
                            return 'NLB'
                    return None
                df['kind'] = df.apply(_kind, axis=1)
                # Replace type for ELBs directly with ALB/NLB
                def _mapped_type(row):
                    if row.get('type') == 'ELBv2LoadBalancer':
                        if row.get('lb_type') == 'application':
                            return 'ALB'
                        if row.get('lb_type') == 'network':
                            return 'NLB'
                        return 'ELBv2'
                    return row.get('type')
                df['type'] = df.apply(_mapped_type, axis=1)
                # Drop helper columns if present
                for helper in ['lb_type', 'kind']:
                    if helper in df.columns:
                        df.drop(columns=[helper], inplace=True)
            df.to_csv(csv_file, index=False, sep=self.config.output.csv_delimiter)
            logger.info(f"Compute resources CSV saved to {csv_file}")
        
        # Network resources
        if inventory.get('network_resources'):
            csv_file = Path(output_dir) / "network_resources.csv"
            df = pd.DataFrame(inventory['network_resources'])
            for col in ['profile', 'account_id', 'arn']:
                if col not in df.columns:
                    df[col] = None
            df.to_csv(csv_file, index=False, sep=self.config.output.csv_delimiter)
            logger.info(f"Network resources CSV saved to {csv_file}")
        
        # Security resources
        if inventory.get('security_resources'):
            csv_file = Path(output_dir) / "security_resources.csv"
            df = pd.DataFrame(inventory['security_resources'])
            for col in ['profile', 'account_id', 'arn']:
                if col not in df.columns:
                    df[col] = None
            df.to_csv(csv_file, index=False, sep=self.config.output.csv_delimiter)
            logger.info(f"Security resources CSV saved to {csv_file}")
    
    def _generate_security_analysis_csv(self, security_analysis: Dict[str, Any], output_dir: str) -> None:
        """Generate CSV for security analysis"""
        
        # Overly permissive security groups
        if security_analysis.get('overly_permissive_sgs'):
            csv_file = Path(output_dir) / "overly_permissive_security_groups.csv"
            df = pd.DataFrame(security_analysis['overly_permissive_sgs'])
            df.to_csv(csv_file, index=False, sep=self.config.output.csv_delimiter)
            logger.info(f"Overly permissive SGs CSV saved to {csv_file}")
        
        # Unused security groups
        if security_analysis.get('unused_sgs'):
            csv_file = Path(output_dir) / "unused_security_groups.csv"
            df = pd.DataFrame(security_analysis['unused_sgs'])
            df.to_csv(csv_file, index=False, sep=self.config.output.csv_delimiter)
            logger.info(f"Unused SGs CSV saved to {csv_file}")
    
    def _generate_excel_report(self, analysis_results: Dict[str, Any], output_dir: str) -> None:
        """Generate comprehensive Excel report with multiple sheets"""
        excel_file = Path(output_dir) / "network_analysis_report.xlsx"
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            
            # Summary sheet
            self._create_summary_sheet(analysis_results, writer)
            
            # Communication paths sheet
            if analysis_results.get('communication_paths'):
                self._create_communication_paths_sheet(analysis_results['communication_paths'], writer)
            
            # Resource inventory sheets
            if analysis_results.get('resource_inventory'):
                self._create_resource_inventory_sheets(analysis_results['resource_inventory'], writer)
            
            # Security analysis sheets
            if analysis_results.get('security_analysis'):
                self._create_security_analysis_sheets(analysis_results['security_analysis'], writer)
            
            # Network insights sheet
            if analysis_results.get('network_insights'):
                self._create_network_insights_sheet(analysis_results['network_insights'], writer)
        
        # Apply formatting
        self._format_excel_workbook(excel_file)
        
        logger.info(f"Excel report saved to {excel_file}")
    
    def _create_summary_sheet(self, analysis_results: Dict[str, Any], writer) -> None:
        """Create summary sheet in Excel"""
        summary_data = []
        
        # Metadata
        metadata = analysis_results.get('metadata', {})
        summary_data.append(['Analysis Timestamp', metadata.get('analysis_timestamp', 'Unknown')])
        summary_data.append(['Total Communication Paths', metadata.get('total_paths_found', 0)])
        summary_data.append(['Cross-Account Paths', metadata.get('cross_account_paths', 0)])
        summary_data.append(['Cross-Region Paths', metadata.get('cross_region_paths', 0)])
        summary_data.append(['Third-Party Paths', metadata.get('third_party_paths', 0)])
        
        # Resource counts
        inventory = analysis_results.get('resource_inventory', {})
        summary = inventory.get('summary', {})
        summary_data.append(['', ''])  # Empty row
        summary_data.append(['Resource Summary', ''])
        summary_data.append(['Total Compute Resources', summary.get('total_compute_resources', 0)])
        summary_data.append(['Total Network Resources', summary.get('total_network_resources', 0)])
        summary_data.append(['Total Security Resources', summary.get('total_security_resources', 0)])
        summary_data.append(['Regions Analyzed', ', '.join(summary.get('regions', []))])
        summary_data.append(['Accounts Analyzed', ', '.join(summary.get('accounts', []))])
        
        # Security insights
        security = analysis_results.get('security_analysis', {})
        summary_data.append(['', ''])  # Empty row
        summary_data.append(['Security Analysis', ''])
        summary_data.append(['Overly Permissive SGs', len(security.get('overly_permissive_sgs', []))])
        summary_data.append(['Unused Security Groups', len(security.get('unused_sgs', []))])
        
        df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_communication_paths_sheet(self, paths: List[Dict], writer) -> None:
        """Create communication paths sheet in Excel"""
        # Flatten path data (same as CSV generation)
        flattened_paths = []
        for path in paths:
            flat_path = {
                'Source_Type': path['source']['type'],
                'Source_ID': path['source']['id'],
                'Source_Name': path['source']['name'],
                'Source_Region': path['source']['region'],
                'Source_Profile': path['source'].get('profile'),
                'Source_Account': path['source'].get('account_id'),
                'Source_Arn': path['source'].get('arn'),
                'Destination_Type': path['destination']['type'],
                'Destination_ID': path['destination']['id'],
                'Destination_Name': path['destination']['name'],
                'Destination_Region': path['destination']['region'],
                'Destination_Profile': path['destination'].get('profile'),
                'Destination_Account': path['destination'].get('account_id'),
                'Destination_Arn': path['destination'].get('arn'),
                'Protocol': path['protocol'],
                'Port_Range': path['port_range'],
                'Direction': path.get('direction'),
                'Is_Cross_Account': path['is_cross_account'],
                'Is_Cross_Region': path['is_cross_region'],
                'Confidence_Score': path['confidence_score'],
            }
            flattened_paths.append(flat_path)
        
        df = pd.DataFrame(flattened_paths)
        df.to_excel(writer, sheet_name='Communication Paths', index=False)
    
    def _create_resource_inventory_sheets(self, inventory: Dict[str, Any], writer) -> None:
        """Create resource inventory sheets in Excel"""
        
        if inventory.get('compute_resources'):
            df = pd.DataFrame(inventory['compute_resources'])
            # Ensure common columns are present
            for col in ['profile', 'account_id', 'arn']:
                if col not in df.columns:
                    df[col] = None
            # Clarify ELB type naming by adding 'kind' column (ALB/NLB)
            if 'type' in df.columns:
                def _kind(row):
                    if row.get('type') == 'ELBv2LoadBalancer':
                        t = row.get('lb_type')
                        if t == 'application':
                            return 'ALB'
                        if t == 'network':
                            return 'NLB'
                    return None
                df['kind'] = df.apply(_kind, axis=1)
                # Replace type for ELBs directly with ALB/NLB
                def _mapped_type(row):
                    if row.get('type') == 'ELBv2LoadBalancer':
                        if row.get('lb_type') == 'application':
                            return 'ALB'
                        if row.get('lb_type') == 'network':
                            return 'NLB'
                        return 'ELBv2'
                    return row.get('type')
                df['type'] = df.apply(_mapped_type, axis=1)
                # Drop helper columns if present
                for helper in ['lb_type', 'kind']:
                    if helper in df.columns:
                        df.drop(columns=[helper], inplace=True)
            df.to_excel(writer, sheet_name='Compute Resources', index=False)
        
        if inventory.get('network_resources'):
            df = pd.DataFrame(inventory['network_resources'])
            for col in ['profile', 'account_id', 'arn']:
                if col not in df.columns:
                    df[col] = None
            df.to_excel(writer, sheet_name='Network Resources', index=False)
        
        if inventory.get('security_resources'):
            df = pd.DataFrame(inventory['security_resources'])
            for col in ['profile', 'account_id', 'arn']:
                if col not in df.columns:
                    df[col] = None
            df.to_excel(writer, sheet_name='Security Resources', index=False)
    
    def _create_security_analysis_sheets(self, security_analysis: Dict[str, Any], writer) -> None:
        """Create security analysis sheets in Excel"""
        
        if security_analysis.get('overly_permissive_sgs'):
            df = pd.DataFrame(security_analysis['overly_permissive_sgs'])
            df.to_excel(writer, sheet_name='Permissive SGs', index=False)
        
        if security_analysis.get('unused_sgs'):
            df = pd.DataFrame(security_analysis['unused_sgs'])
            df.to_excel(writer, sheet_name='Unused SGs', index=False)
    
    def _create_network_insights_sheet(self, insights: Dict[str, Any], writer) -> None:
        """Create network insights sheet in Excel"""
        insights_data = []
        
        # Connectivity summary
        connectivity = insights.get('connectivity_summary', {})
        insights_data.append(['Connectivity Summary', ''])
        for key, value in connectivity.items():
            insights_data.append([key.replace('_', ' ').title(), value])
        
        # Security recommendations
        recommendations = insights.get('security_recommendations', [])
        if recommendations:
            insights_data.append(['', ''])
            insights_data.append(['Security Recommendations', ''])
            for rec in recommendations:
                insights_data.append([rec.get('type', ''), rec.get('description', '')])
        
        df = pd.DataFrame(insights_data, columns=['Category', 'Details'])
        df.to_excel(writer, sheet_name='Network Insights', index=False)
    
    def _format_excel_workbook(self, excel_file: Path) -> None:
        """Apply formatting to Excel workbook"""
        try:
            wb = openpyxl.load_workbook(excel_file)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center")
            
            # Format each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Format header row
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                
                # Freeze the first row and enable auto-filters
                try:
                    ws.freeze_panes = "A2"
                    # Apply auto filter to the used range
                    ws.auto_filter.ref = ws.dimensions
                except Exception:
                    pass
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_file)
            
        except Exception as e:
            logger.warning(f"Failed to format Excel workbook: {str(e)}")
    
    def _generate_html_report(self, analysis_results: Dict[str, Any], output_dir: str) -> None:
        """Generate interactive HTML report"""
        html_file = Path(output_dir) / "network_analysis_report.html"
        
        # Create HTML template
        html_template = self._get_html_template()
        
        # Prepare data for template
        template_data = {
            'title': 'AWS Network Analysis Report',
            'metadata': analysis_results.get('metadata', {}),
            'summary': self._prepare_summary_data(analysis_results),
            'communication_paths': analysis_results.get('communication_paths', []),
            'resource_inventory': analysis_results.get('resource_inventory', {}),
            'security_analysis': analysis_results.get('security_analysis', {}),
            'network_insights': analysis_results.get('network_insights', {}),
            'compliance_report': analysis_results.get('compliance_report', {}),
        }
        
        # Render template
        template = Template(html_template)
        html_content = template.render(**template_data)
        
        # Save HTML file
        with open(html_file, 'w') as f:
            f.write(html_content)
        
        logger.info(f"HTML report saved to {html_file}")
    
    def _prepare_summary_data(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare summary data for HTML template"""
        metadata = analysis_results.get('metadata', {})
        inventory = analysis_results.get('resource_inventory', {})
        summary = inventory.get('summary', {})
        
        return {
            'total_paths': metadata.get('total_paths_found', 0),
            'cross_account_paths': metadata.get('cross_account_paths', 0),
            'cross_region_paths': metadata.get('cross_region_paths', 0),
            'compute_resources': summary.get('total_compute_resources', 0),
            'network_resources': summary.get('total_network_resources', 0),
            'security_resources': summary.get('total_security_resources', 0),
            'regions': summary.get('regions', []),
            'accounts': summary.get('accounts', []),
        }
    
    def _get_html_template(self) -> str:
        """Get HTML template for the report"""
        return """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ title }}</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        h1, h2, h3 {
            color: #333;
        }
        h1 {
            text-align: center;
            border-bottom: 3px solid #366092;
            padding-bottom: 10px;
        }
        .summary-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin: 30px 0;
        }
        .summary-card {
            background: linear-gradient(135deg, #366092, #4a7ba7);
            color: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }
        .summary-card h3 {
            margin: 0 0 10px 0;
            color: white;
        }
        .summary-card .number {
            font-size: 2em;
            font-weight: bold;
        }
        .table-container {
            overflow-x: auto;
            margin: 20px 0;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }
        th, td {
            border: 1px solid #ddd;
            padding: 12px;
            text-align: left;
        }
        th {
            background-color: #366092;
            color: white;
            font-weight: bold;
        }
        tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        tr:hover {
            background-color: #f5f5f5;
        }
        .filter-container {
            margin: 20px 0;
            padding: 15px;
            background-color: #f8f9fa;
            border-radius: 5px;
        }
        .filter-input {
            padding: 8px;
            margin: 5px;
            border: 1px solid #ddd;
            border-radius: 4px;
        }
        .section {
            margin: 40px 0;
            padding: 20px;
            border-left: 4px solid #366092;
            background-color: #fafafa;
        }
        .badge {
            display: inline-block;
            padding: 4px 8px;
            border-radius: 4px;
            font-size: 0.8em;
            font-weight: bold;
        }
        .badge-success { background-color: #28a745; color: white; }
        .badge-warning { background-color: #ffc107; color: black; }
        .badge-danger { background-color: #dc3545; color: white; }
        .badge-info { background-color: #17a2b8; color: white; }
    </style>
    <script>
        function filterTable(tableId, inputId) {
            const input = document.getElementById(inputId);
            const filter = input.value.toUpperCase();
            const table = document.getElementById(tableId);
            const tr = table.getElementsByTagName("tr");
            
            for (let i = 1; i < tr.length; i++) {
                let td = tr[i].getElementsByTagName("td");
                let display = false;
                
                for (let j = 0; j < td.length; j++) {
                    if (td[j]) {
                        let txtValue = td[j].textContent || td[j].innerText;
                        if (txtValue.toUpperCase().indexOf(filter) > -1) {
                            display = true;
                            break;
                        }
                    }
                }
                
                tr[i].style.display = display ? "" : "none";
            }
        }
        
        function sortTable(tableId, columnIndex) {
            const table = document.getElementById(tableId);
            let switching = true;
            let dir = "asc";
            let switchcount = 0;
            
            while (switching) {
                switching = false;
                const rows = table.rows;
                
                for (let i = 1; i < (rows.length - 1); i++) {
                    let shouldSwitch = false;
                    const x = rows[i].getElementsByTagName("TD")[columnIndex];
                    const y = rows[i + 1].getElementsByTagName("TD")[columnIndex];
                    
                    if (dir == "asc") {
                        if (x.innerHTML.toLowerCase() > y.innerHTML.toLowerCase()) {
                            shouldSwitch = true;
                            break;
                        }
                    } else if (dir == "desc") {
                        if (x.innerHTML.toLowerCase() < y.innerHTML.toLowerCase()) {
                            shouldSwitch = true;
                            break;
                        }
                    }
                }
                
                if (shouldSwitch) {
                    rows[i].parentNode.insertBefore(rows[i + 1], rows[i]);
                    switching = true;
                    switchcount++;
                } else {
                    if (switchcount == 0 && dir == "asc") {
                        dir = "desc";
                        switching = true;
                    }
                }
            }
        }
    </script>
</head>
<body>
    <div class="container">
        <h1>{{ title }}</h1>
        
        <!-- Summary Section -->
        <div class="section">
            <h2>Executive Summary</h2>
            <div class="summary-grid">
                <div class="summary-card">
                    <h3>Communication Paths</h3>
                    <div class="number">{{ summary.total_paths }}</div>
                </div>
        
        <!-- Resource Inventory Section -->
        <div class="section">
            <h2>Resource Inventory</h2>
            
            {% if resource_inventory.compute_resources %}
            <h3>Compute Resources</h3>
            <div class="table-container">
                <table>
                    <thead>
                        <tr>
                            <th>Type</th>
                            <th>Name/ID</th>
                            <th>Region</th>
                            <th>Profile</th>
                            <th>Account</th>
                            <th>ARN</th>
                            <th>VPC</th>
                            <th>Subnets</th>
                            <th>Security Groups</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for res in resource_inventory.compute_resources %}
                        <tr>
                            <td>
                                {{ res.type }}
                                {% if res.lb_type %}
                                    ({{ 'ALB' if res.lb_type == 'application' else ('NLB' if res.lb_type == 'network' else res.lb_type) }})
                                {% endif %}
                            </td>
                            <td>{{ res.name or res.id }}</td>
                            <td>{{ res.region }}</td>
                            <td>{{ res.profile }}</td>
                            <td>{{ res.account_id }}</td>
                            <td>{{ res.arn }}</td>
                            <td>{{ res.vpc_id }}</td>
                            <td>
                                {% if res.subnet_id %}{{ res.subnet_id }}{% endif %}
                                {% if res.subnet_ids %}{{ res.subnet_ids | join(', ') }}{% endif %}
                            </td>
                            <td>
                                {% if res.security_groups %}
                                    {% if res.security_groups[0] is string %}
                                        {{ res.security_groups | join(', ') }}
                                    {% else %}
                                        {{ res.security_groups | map(attribute='GroupId') | list | join(', ') }}
                                    {% endif %}
                                {% endif %}
                            </td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
            {% endif %}
            
            {% if resource_inventory.network_resources %}
            <h3>Network Resources</h3>
            <div class="table-container">
                <table>
                    <thead>
                        <tr>
                            <th>Type</th>
                            <th>Name/ID</th>
                            <th>Region</th>
                            <th>Profile</th>
                            <th>Account</th>
                            <th>ARN</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for res in resource_inventory.network_resources %}
                        <tr>
                            <td>{{ res.type }}</td>
                            <td>{{ res.name or res.id }}</td>
                            <td>{{ res.region }}</td>
                            <td>{{ res.profile }}</td>
                            <td>{{ res.account_id }}</td>
                            <td>{{ res.arn }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
            {% endif %}
            
            {% if resource_inventory.security_resources %}
            <h3>Security Resources</h3>
            <div class="table-container">
                <table>
                    <thead>
                        <tr>
                            <th>Type</th>
                            <th>Name/ID</th>
                            <th>Region</th>
                            <th>Profile</th>
                            <th>Account</th>
                            <th>ARN</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for res in resource_inventory.security_resources %}
                        <tr>
                            <td>{{ res.type }}</td>
                            <td>{{ res.name or res.id }}</td>
                            <td>{{ res.region }}</td>
                            <td>{{ res.profile }}</td>
                            <td>{{ res.account_id }}</td>
                            <td>{{ res.arn }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
            {% endif %}
        </div>
                <div class="summary-card">
                    <h3>Compute Resources</h3>
                    <div class="number">{{ summary.compute_resources }}</div>
                </div>
                <div class="summary-card">
                    <h3>Network Resources</h3>
                    <div class="number">{{ summary.network_resources }}</div>
                </div>
                <div class="summary-card">
                    <h3>Security Resources</h3>
                    <div class="number">{{ summary.security_resources }}</div>
                </div>
                <div class="summary-card">
                    <h3>Cross-Account Paths</h3>
                    <div class="number">{{ summary.cross_account_paths }}</div>
                </div>
                <div class="summary-card">
                    <h3>Cross-Region Paths</h3>
                    <div class="number">{{ summary.cross_region_paths }}</div>
                </div>
            </div>
        </div>
        
        <!-- Communication Paths Section -->
        <div class="section">
            <h2>Communication Paths</h2>
            <div class="filter-container">
                <input type="text" id="pathFilter" class="filter-input" placeholder="Filter communication paths..." 
                       onkeyup="filterTable('pathsTable', 'pathFilter')">
            </div>
            <div class="table-container">
                <table id="pathsTable">
                    <thead>
                        <tr>
                            <th onclick="sortTable('pathsTable', 0)">Source</th>
                            <th onclick="sortTable('pathsTable', 1)">Source Profile</th>
                            <th onclick="sortTable('pathsTable', 2)">Source Account</th>
                            <th onclick="sortTable('pathsTable', 3)">Destination</th>
                            <th onclick="sortTable('pathsTable', 4)">Destination Profile</th>
                            <th onclick="sortTable('pathsTable', 5)">Destination Account</th>
                            <th onclick="sortTable('pathsTable', 6)">Protocol</th>
                            <th onclick="sortTable('pathsTable', 7)">Port</th>
                            <th onclick="sortTable('pathsTable', 8)">Cross-Account</th>
                            <th onclick="sortTable('pathsTable', 9)">Cross-Region</th>
                            <th onclick="sortTable('pathsTable', 10)">Confidence</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for path in communication_paths %}
                        <tr>
                            <td>{{ path.source.name }} ({{ path.source.type }})</td>
                            <td>{{ path.source.profile }}</td>
                            <td>{{ path.source.account_id }}</td>
                            <td>{{ path.destination.name }} ({{ path.destination.type }})</td>
                            <td>{{ path.destination.profile }}</td>
                            <td>{{ path.destination.account_id }}</td>
                            <td>{{ path.protocol }}</td>
                            <td>{{ path.port_range }}</td>
                            <td>
                                {% if path.is_cross_account %}
                                    <span class="badge badge-warning">Yes</span>
                                {% else %}
                                    <span class="badge badge-success">No</span>
                                {% endif %}
                            </td>
                            <td>
                                {% if path.is_cross_region %}
                                    <span class="badge badge-info">Yes</span>
                                {% else %}
                                    <span class="badge badge-success">No</span>
                                {% endif %}
                            </td>
                            <td>{{ "%.2f"|format(path.confidence_score) }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        </div>
        
        <!-- Security Analysis Section -->
        <div class="section">
            <h2>Security Analysis</h2>
            {% if security_analysis.overly_permissive_sgs %}
            <h3>Overly Permissive Security Groups</h3>
            <div class="table-container">
                <table>
                    <thead>
                        <tr>
                            <th>Security Group ID</th>
                            <th>Name</th>
                            <th>Region</th>
                            <th>Issues</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for sg in security_analysis.overly_permissive_sgs %}
                        <tr>
                            <td>{{ sg.sg_id }}</td>
                            <td>{{ sg.sg_name }}</td>
                            <td>{{ sg.region }}</td>
                            <td>
                                {% for issue in sg.issues %}
                                    <span class="badge badge-danger">{{ issue }}</span>
                                {% endfor %}
                            </td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
            {% endif %}
        </div>
        
        <!-- Network Insights Section -->
        <div class="section">
            <h2>Network Insights</h2>
            {% if network_insights.security_recommendations %}
            <h3>Security Recommendations</h3>
            <ul>
                {% for rec in network_insights.security_recommendations %}
                <li><strong>{{ rec.type|title }}:</strong> {{ rec.description }}</li>
                {% endfor %}
            </ul>
            {% endif %}
            
            {% if network_insights.optimization_opportunities %}
            <h3>Optimization Opportunities</h3>
            <ul>
                {% for opp in network_insights.optimization_opportunities %}
                <li><strong>{{ opp.type|title }}:</strong> {{ opp.description }}</li>
                {% endfor %}
            </ul>
            {% endif %}
        </div>
        
        <footer style="text-align: center; margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; color: #666;">
            <p>Generated by AWS Network Discovery and Analysis Tool</p>
            <p>Report generated at: {{ metadata.analysis_timestamp }}</p>
        </footer>
    </div>
</body>
</html>
        """
